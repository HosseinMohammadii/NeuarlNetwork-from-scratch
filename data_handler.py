from openpyxl import load_workbook, Workbook
import numpy as np


def target_converter(target):
    if target == 'Iris-setosa':
        return 0.16667
    if target == 'Iris-versicolor':
        return 0.5
    if target == 'Iris-virginica':
        return 0.83337


def read_data(input_file_path):
    data = []
    wb = load_workbook(input_file_path)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for i in range(1, sheet.max_row):
            d = [sheet.cell(i, j).value for j in range(1, 5)]
            d = np.array([d]).T
            t = np.array([[target_converter(sheet.cell(i, 5).value)]])
            data.append((d, t))
    return data
